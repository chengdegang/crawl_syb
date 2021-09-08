import os
import zipfile

from openpyxl.workbook import Workbook
import openpyxl
import time
from openpyxl.cell import MergedCell

file1 = ''
file_xxx2 = '/Users/jackrechard/PycharmProjects/testexcel/file/xxx2.xlsx'

def write_excel(file = '默认.xlsx',data = [['默认数据3','默认数据4'],['默认数据5','默认数据6']],
                sheetname = 'ces',creat = True):
    """
    在输入路径创建xlsx文件,若已存在会直接覆盖写入数据，需要重写为若输入路径不存在文件，则在输入路径下创建文件并写入
    """
    if os.path.exists(file) == True:
        wb = openpyxl.load_workbook(file)
        # 默认写到第一个sheet中，index为0
        wb.create_sheet(index=0, title=f'{sheetname}')
        sheet = wb[f'{sheetname}']
        # 好的写法写入excel
        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        wb.save(file)
        print('---写入完成---')
    if creat == True:
        wbc = Workbook()
        # path = file
        # path = file.split('/')[:-1]
        wbc.save(file)
        wb = openpyxl.load_workbook(file)
        # 默认写到第一个sheet中，index为0
        wb.create_sheet(index=0, title=f'{sheetname}')
        sheet = wb[f'{sheetname}']
        # 好的写法写入excel
        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        wb.save(file)
        print('---写入完成---')
        return file
    else:
        wbc = Workbook()
        log_path = os.getcwd() + '/'
        # t = time.strftime('%Y%m%d_%H%M', time.localtime(time.time()))
        # suffix = '.xlsx'  # 文件类型
        # newfile = file + suffix
        path = log_path + file
        wbc.save(path)
        print(f"创建文件 {log_path + file} ")
        wb = openpyxl.load_workbook(log_path + file)
        # 默认写到第一个sheet中，index为0
        wb.create_sheet(index=0, title=f'{sheetname}')
        sheet = wb[f'{sheetname}']
        # 好的写法写入excel
        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        wb.save(log_path + file)
        print('---写入完成---')
        return log_path + file

"""
支持xlsx格式的读,输入要读的文件路径,筛选需要的数据
"""
def openpy_read_xlsx(file,sheetname):
    wb = openpyxl.load_workbook(file)
    table = wb[sheetname]
    #获取行数列数
    nrow = table.max_row
    ncol = table.max_column
    need_data2 = []

    #获取标题
    for row in range(1,nrow+1):
        for col in range(1,ncol+1):
            #如果这一行中存在title字样，遍历这一行的所有数据，添加到列表中
            if "序号" in str(table.cell(row, col).value):
                for i in range(1,ncol+1):
                    # print(table.cell(row, i).value)
                    need_data2.append(table.cell(row, i).value)
                # print(need_data2)
                break

    #获取需要的行，判断条件为包含'_1'
    for row in range(1,nrow+1):
        for col in range(1,ncol+1):
            #判断是否包含'value'，是的话选取整列
            if "专业" in str(table.cell(row,col).value):
                #判断这列是否包含'_1'，是的话取整行
                # print(table.cell(row, col).value)
                #遍历这一列的所有数据，并标记存在'_1'的行
                for i in range(1,nrow+1):
                    if "情报学" in str(table.cell(i,col).value):
                        # print(table.cell(i,col).value
                        #将找到的这一行所有数据写入列表
                        for j in range(1,ncol+1):
                            # print(table.cell(i,j).value)
                            need_data2.append(table.cell(i,j).value)

    #将该列表按照列数切成多个列表
    per_list_len = ncol
    list_of_group = zip(*(iter(need_data2),) * per_list_len)
    end_list = [list(i) for i in list_of_group]  # i is a tuple
    count = len(need_data2) % per_list_len
    end_list.append(need_data2[-count:]) if count != 0 else end_list
    # print(end_list)
    return end_list

def merged_deal_xlsx(file):
    """
    处理xlsx的excel合并单元格,获取文件路径，重读excel文件并输出每一个单元格重写了数据的列表（主要针对有合并单元格的情况），并去重
    """
    newdata = []
    wb = openpyxl.load_workbook(file)
    # table = wb["Sheet2"]
    sheetnames = wb.sheetnames #
    table = wb[sheetnames[0]] #获取第一个sheet
    nrow = table.max_row
    ncol = table.max_column

    for row_index in range(1, nrow + 1):
        for col_index in range(1, ncol + 1):
            cell = table.cell(row=row_index, column=col_index)
            if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
                for merged_range in table.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                    if cell.coordinate in merged_range:
                        # 获取合并区域左上角的单元格作为该单元格的值返回
                        cell_ = table.cell(row=merged_range.min_row, column=merged_range.min_col)
                        newdata.append(cell_.value)
                        break
            else:
                cell_ = table.cell(row=row_index, column=col_index)
                # newdata.append(cell_.value)
                newdata.append(cell_.value)

    per_list_len = ncol
    list_of_group = zip(*(iter(newdata),) * per_list_len)
    end_list = [list(i) for i in list_of_group]  # i is a tuple
    count = len(newdata) % per_list_len
    end_list.append(newdata[-count:]) if count != 0 else end_list

    #去重
    end_list2 = []
    for element in end_list:
        if element not in end_list2:
            end_list2.append(element)
    # print(end_list2)
    return end_list2

#第一步，读取xlsx的file文件，并将其合并补全每一个合并单元格的数据，返回list，并写入到xlsx文件
#第二步，读取第一步保存的文件并筛选指定的数据返回一个list
#第三步，将这个list数据写入到文件中，命名一个sheetname
if __name__ == '__main__':
    # # write_excel(file=file_xxx2, data=merged_deal_xlsx(file_xxx2), sheetname='tempxlsx')
    # # data = openpy_read_xlsx(file=file_xxx2, sheetname='tempxlsx')
    # # write_excel(file=file_xxx2, data=data, sheetname='finalxlsx')
    #
    # filesz = '/Users/jackrechard/PycharmProjects/crawl_syb/download/2021年浙江杭州电子科技大学招聘公告（第四批）.xlsx'
    # # filesz = '/Users/jackrechard/PycharmProjects/crawl_syb/download/2021年浙江工商大学招聘公告（第三批）.xlsx' #实战
    # # filesz = '/Users/jackrechard/PycharmProjects/crawl_syb/download/2021年浙江师范大学招聘公告（第三批）.xlsx'
    # ff = filesz
    # try :
    #     write_excel(file=ff, data=merged_deal_xlsx(ff), sheetname='tempxlsx')
    #     data = openpy_read_xlsx(file=ff, sheetname='tempxlsx')
    #     # print(len(data)) #无数据时等于1
    #     write_excel(file=ff, data=data, sheetname='finalxlsx')
    # except zipfile.BadZipFile:
    #     print(f'文件异常，请检查,路径为：{ff}')
    write_excel('/Users/jackrechard/Desktop/ces666.xlsx')

